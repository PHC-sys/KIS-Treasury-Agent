# -*- coding: utf-8 -*-
"""
infomax_pull.py — 인포맥스 애드인 엑셀 → ledger 리더/수집기

인포맥스 애드인(IMDH 히스토리 함수)으로 채운 엑셀을 읽어 ledger에 적재한다.
애드인은 그 PC의 Excel + 인포맥스 로그인 세션에서만 값이 산다.

데이터는 통합 워크북 infomax_data.xlsx(시트 7개, 필요한 항목만). build_infomax.py로 생성.

모드:
  dump    "경로.xlsx"  — 엑셀 시트/셀 레이아웃 출력 (캐시값, Excel 불필요)
  refresh [경로]       — COM 재조회. 경로 없으면 infomax_data.xlsx(시트 전체) 갱신
  load                 — infomax_data.xlsx → sanity → ledger 적재 → 파생계산 → export

일일 자동화 순서: refresh → load  (run.py가 update.py와 함께 호출)
LOAD_SPEC = 인포맥스 시트·항목 매핑 / build_infomax.py의 SHEETS = 종목 정의(단일 진실).
"""
import sys
from pathlib import Path


def dump(path):
    """저장된 엑셀의 모든 시트 used range를 출력. IMDH 캐시값을 그대로 보여준다."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)   # data_only = 캐시된 계산값
    print(f"파일: {path}")
    print(f"시트: {wb.sheetnames}\n")
    for ws in wb.worksheets:
        print("=" * 60)
        print(f"[{ws.title}]  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
        # 앞 12행만 미리보기 (레이아웃 파악엔 충분)
        for r, row in enumerate(ws.iter_rows(values_only=True), 1):
            if r > 12:
                print("   ... (이하 생략)")
                break
            cells = ["" if c is None else str(c) for c in row]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                print(f"  r{r:<3} " + " | ".join(cells))


# ── 읽기 스펙: 생성된 엑셀 → 우리 field 매핑 ──────────────────────────
# IMDH 출력: '일자' 헤더행 아래로 데이터. 값 = 지정한 '항목 헤더 텍스트'의 열.
# 통합 워크북 infomax_data.xlsx (시트별 IMDH, 필요한 항목만). build_infomax.py로 생성.
# 각 항목: sheet, header_text(날짜열=일자, row3), fields{우리필드: 항목헤더 or ("net",매수,매도)}
# 프로젝트 루트(=src/의 부모) 기준 절대경로 → 실행 cwd와 무관하게 워크북을 찾음.
IMX_FILE = str(Path(__file__).resolve().parent.parent / "infomax_data.xlsx")


def _fut_fields(suf):
    """FUT 시트(연결선물) → 우리 필드 매핑. 순매수 = 매수수량−매도수량."""
    return {
        f"ktb{suf}_settle": "현재가",           # 정산가는 옛 데이터 0 많음 → 현재가(종가)
        f"ktb{suf}_chg": "전일대비",
        f"ktb{suf}_open": "시가",
        f"ktb{suf}_high": "고가",
        f"ktb{suf}_low": "저가",
        f"ktb{suf}_vol": "거래량",
        f"ktb{suf}_oi": "미결제약정수량",
        f"ktb{suf}_theo_basis": "이론베이시스",
        f"fx_net_ktb{suf}": ("net", "외국인합매수수량", "외국인합매도수량"),
        f"bank_net_ktb{suf}": ("net", "은행매수수량", "은행매도수량"),
        f"itrust_net_ktb{suf}": ("net", "투신매수수량", "투신매도수량"),
        f"ins_net_ktb{suf}": ("net", "보험매수수량", "보험매도수량"),
        # v2.6 — 인포맥스가 순매수수량을 직접 주므로 그대로 받는다.
        f"sec_net_ktb{suf}": "증권/선물순매수수량",
        f"indiv_net_ktb{suf}": "개인순매수수량",
        f"pension_net_ktb{suf}": "연기금순매수수량",
        # etc = "나머지 전부의 합"(차장님 정의). 초기 구간엔 없는 항목이 있어 빈칸=0으로 더한다.
        f"etc_net_ktb{suf}": ("sum", "기타순매수수량", "국가/지방순매수수량",
                              "연기금순매수수량", "종신금순매수수량"),
    }


LOAD_SPEC = [
    {"sheet": "FUT3",   "header_text": "일자", "fields": _fut_fields("3")},
    {"sheet": "FUT10",  "header_text": "일자", "fields": _fut_fields("10")},
    {"sheet": "IRS3Y",  "header_text": "일자", "fields": {"irs3y": "MID종가"}},
    {"sheet": "IRS10Y", "header_text": "일자", "fields": {"irs10y": "MID종가"}},
    {"sheet": "OIS3Y",  "header_text": "일자", "fields": {"ois3y": "MID종가"}},
    {"sheet": "KOFR",   "header_text": "일자", "fields": {"kofr": "현재가"}},  # ECO/785831 (ECOS보다 신선)
]
for _s in LOAD_SPEC:            # 전부 같은 통합 파일
    _s["file"] = IMX_FILE

# MACRO 5종: 우리필드 → 워크북 시트명 (계산·배치는 config.MACRO_SPEC + _read_macro)
MACRO_SHEETS = {
    "cpi_yoy": "M_CPI", "core_cpi_yoy": "M_CORE", "bei": "M_BEI",
    "exports_yoy": "M_EXP", "ip_yoy": "M_IP",
}


def _to_date(v):
    import datetime as dt
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()[:10].replace("/", "-").replace(".", "-")
    return s if len(s) == 10 else None


def _read_one(spec):
    """LOAD_SPEC 항목 하나 → Record 리스트 (헤더 텍스트로 열 매칭)."""
    import openpyxl
    from collectors.base import Record
    ws = openpyxl.load_workbook(spec["file"], data_only=True)[spec["sheet"]]
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows)
              if r and spec["header_text"] in [str(c).strip() if c else "" for c in r])
    header = [str(c).strip() if c else "" for c in rows[hi]]
    dcol = header.index(spec["header_text"])

    # ★워크북에 없는 항목은 None으로 돌려준다(예외 아님).
    #  회사 zip이 확장 전 워크북으로 덮어쓰면 header.index가 ValueError를 던지는데,
    #  load()는 FileNotFoundError/KeyError만 잡아서 시트 전체 적재가 통째로 죽었다
    #  (신규 컬럼만 비는 게 아니라 그날 선물 데이터가 전부 안 들어옴).
    #  → 없는 항목만 조용히 빠지고 나머지는 정상 적재. 단 무엇이 빠졌는지는 아래서 경고한다.
    missing = sorted({n for f in spec["fields"].values()
                      for n in ((f[1:] if isinstance(f, tuple) else (f,)))
                      if n not in header})
    if missing:
        print(f"  ⚠ [{spec['sheet']}] 워크북에 없는 항목 {len(missing)}개 → 해당 필드 건너뜀: "
              f"{', '.join(missing)}  (build_infomax.py로 워크북 갱신 필요)")

    def cell(r, name):
        if name not in header:
            return None
        c = header.index(name)
        return r[c] if c < len(r) else None

    recs = []
    for our_field, item in spec["fields"].items():
        for r in rows[hi + 1:]:
            d = _to_date(r[dcol]) if r[dcol] is not None else None
            if not d:
                continue
            if isinstance(item, tuple) and item[0] == "net":   # 매수 − 매도
                buy, sell = cell(r, item[1]), cell(r, item[2])
                if buy in (None, "") or sell in (None, ""):
                    continue
                v = float(buy) - float(sell)
            elif isinstance(item, tuple) and item[0] == "sum":  # 여러 항목 합
                # 빈칸 = 그 분류가 아직 없던 시기 → 0으로 취급. 단 전부 비면 그 날은 값 없음.
                got = [cell(r, name) for name in item[1:]]
                got = [g for g in got if g not in (None, "")]
                if not got:
                    continue
                v = float(sum(float(g) for g in got))
            else:                                              # 직접 열
                raw = cell(r, item)
                if raw in (None, ""):
                    continue
                v = float(raw)
            # 가격필드(settle/OHLC/basis) 정확히 0 = 미기록 → 빈칸(0=진짜 규칙).
            # vol/oi/net의 0은 진짜(거래없음/순매수0)라 유지.
            if v == 0 and our_field.endswith(
                    ("_settle", "_open", "_high", "_low", "_theo_basis")):
                continue
            recs.append(Record(date=d, field=our_field, value=v, as_of=d))
    return recs


def _read_macro():
    """MACRO 5종: 각 시트(일자=참조월말, 현재가) → 지수/금액이면 yoy 계산 →
    발표일(config 규칙)에 배치 → 거래일마다 ffill(다음 발표 전까지 직전값 유지)."""
    import bisect
    import openpyxl
    import config
    from collectors.base import Record

    TD = [d.isoformat() for d in config.trading_days(config.BACKFILL_START, config.ref_date())]

    def ftd(s):                       # s(YYYY-MM-DD) 이후 첫 거래일
        i = bisect.bisect_left(TD, s)
        return TD[i] if i < len(TD) else None

    def addm(ym, n):                  # 'YYYY-MM' + n개월
        y, m = int(ym[:4]), int(ym[5:7])
        m += n; y += (m - 1) // 12; m = (m - 1) % 12 + 1
        return f"{y:04d}-{m:02d}"

    place = {"d5": lambda m: ftd(addm(m, 1) + "-05"), "m1": lambda m: ftd(addm(m, 1) + "-01"),
             "m2": lambda m: ftd(addm(m, 2) + "-01"), "d16": lambda m: ftd(addm(m, 1) + "-16")}

    wb = openpyxl.load_workbook(IMX_FILE, data_only=True)
    recs = []
    for field, (code, kind, pl) in config.MACRO_SPEC.items():
        sheet = MACRO_SHEETS[field]
        if sheet not in wb.sheetnames:
            print(f"  (건너뜀: MACRO 시트 없음 '{sheet}' — build_infomax.py 재실행 필요)")
            continue
        rows = list(wb[sheet].iter_rows(values_only=True))
        hi = next(i for i, r in enumerate(rows)
                  if r and "일자" in [str(c).strip() if c else "" for c in r])
        hdr = [str(c).strip() if c else "" for c in rows[hi]]
        di, vi = hdr.index("일자"), hdr.index("현재가")
        series = {}                   # 참조월(YYYY-MM) → 현재가(지수/금액/값)
        for r in rows[hi + 1:]:
            d = _to_date(r[di]) if di < len(r) and r[di] is not None else None
            if not d or vi >= len(r) or r[vi] in (None, ""):
                continue
            series[d[:7]] = float(r[vi])
        pts = {}                      # 발표일(거래일) → 값
        for m in sorted(series):
            if kind == "yoy":
                py = addm(m, -12)
                if py not in series or series[py] == 0:
                    continue
                val = (series[m] / series[py] - 1) * 100
            else:
                val = series[m]
            pd = place[pl](m)
            if pd:
                pts[pd] = val
        keys = sorted(pts); last = None; ki = 0    # 거래일마다 ffill
        for T in TD:
            while ki < len(keys) and keys[ki] <= T:
                last = pts[keys[ki]]; ki += 1
            if last is not None:
                recs.append(Record(date=T, field=field, value=last, as_of=T))
    return recs


def _asof_records(us_rows, kdays, field):
    """미국일자 시계열 → 한국거래일로 as-of 배정 (ust10·ust2·wti 공용).
    us_rows: [(미국일 'YYYY-MM-DD', 값), ...]. kdays: 정렬된 한국거래일 문자열.
    ★as-of(룩어헤드 방지): 미국일 d의 종가는 한국 다음거래일 새벽 확정 → d를 '첫 한국거래일 > d'에 배정.
    갭>10일이면 범위밖(예: 우리 데이터 이전 과거일)으로 스킵. 0/결측 스킵.

    ★★한 한국일에 미국일이 여러 개 몰리면(한국 휴장 연휴) **가장 최신 미국일**을 채택한다.
      한국 T 개장 시점에 알 수 있는 건 '직전 미국 종가'이므로 최신이 맞다.
      버그 이력(2026-08-18 수정): 예전엔 그냥 append 했는데, store.upsert가 executemany라
      '나중 실행'이 이긴다. 워크북은 sort=D(내림차순)이라 **오래된 미국일이 최종값**이 됐다.
      (폐기된 fred.py는 CSV가 오름차순이라 우연히 맞았고, 인포맥스로 옮기며 전제가 깨졌다.)
      실측 오류: 한국 2026-03-03=미국 02-27(3.942, 정답 03-02 4.035) 등.
      → 입력 순서와 무관하게 동작하도록 한국일별 max(미국일)로 확정한다.
    반환 Record(date=한국일, field, value, as_of=미국일)."""
    import bisect
    from datetime import date as _date
    from collectors.base import Record
    best = {}                                       # 한국일 → (미국일, 값)
    for d, v in us_rows:
        if not d or v in (None, ""):
            continue
        v = float(v)
        if v == 0:                                  # 0 = 미기록 → 빈칸(규율 §6)
            continue
        i = bisect.bisect_right(kdays, d)           # d 다음 한국거래일(첫 T > d)
        if i >= len(kdays):
            continue                                # 아직 배정할 한국거래일 없음(최신 미국일) → 다음 실행에서
        kd = kdays[i]
        if (_date.fromisoformat(kd) - _date.fromisoformat(d)).days > 10:
            continue                                # 큰 갭(범위밖 과거일이 첫 한국일에 몰림) → 스킵
        if kd not in best or d > best[kd][0]:       # 최신 미국일 우선
            best[kd] = (d, v)
    return [Record(date=kd, field=field, value=v, as_of=d)
            for kd, (d, v) in sorted(best.items())]


def _read_asof(sheet_name, fields):
    """워크북 해외 시트(미국일자 + 항목열들) → as-of 시프트 → Records.

    fields: {우리필드: 워크북 항목헤더}  — 한 시트에서 여러 항목을 한 번에 읽는다.
            (예 US10Y: {"ust10":"MID_Close", "ust10_open":"MID_Open", ...})
    워크북/거래일 목록은 시트당 1회만 로드한다.
    없는 항목은 건너뛰고 경고 — 확장 전 워크북으로 덮어써도 죽지 않게.
    """
    import openpyxl
    import config
    if not Path(IMX_FILE).exists():
        return []
    wb = openpyxl.load_workbook(IMX_FILE, data_only=True)
    if sheet_name not in wb.sheetnames:
        print(f"  (건너뜀: {sheet_name} 시트 없음 — build_infomax.py add 또는 run.bat 재실행 필요)")
        return []
    rows = list(wb[sheet_name].iter_rows(values_only=True))
    hi = next((i for i, r in enumerate(rows)
               if r and "일자" in [str(c).strip() if c else "" for c in r]), None)
    if hi is None:
        return []
    hdr = [str(c).strip() if c else "" for c in rows[hi]]
    di = hdr.index("일자")
    missing = [h for h in fields.values() if h not in hdr]
    if missing:
        print(f"  ⚠ [{sheet_name}] 워크북에 없는 항목 → 해당 필드 건너뜀: {', '.join(missing)}")
    kdays = [k.isoformat() for k in
             config.trading_days(config.BACKFILL_START, config.ref_date())]

    out = []
    for field, value_header in fields.items():
        if value_header not in hdr:
            continue
        vi = hdr.index(value_header)
        us_rows = []
        for r in rows[hi + 1:]:
            d = _to_date(r[di]) if di < len(r) and r[di] is not None else None
            v = r[vi] if vi < len(r) else None
            if d:
                us_rows.append((d, v))
        out += _asof_records(us_rows, kdays, field)
    return out


def _read_ust10():
    """US10Y 시트 → 미국채10년 MID OHLC. ust10 = MID_Close(기존 컬럼 유지)."""
    return _read_asof("US10Y", {
        "ust10":      "MID_Close",
        "ust10_open": "MID_Open",
        "ust10_high": "MID_High",
        "ust10_low":  "MID_Low",
    })


def _read_ust2():
    """US02Y 시트 → 미국채2년 MID OHLC. (만기 2자리 제로패딩: US02Y — US2Y는 없음)"""
    return _read_asof("US02Y", {
        "ust2":      "MID_Close",
        "ust2_open": "MID_Open",
        "ust2_high": "MID_High",
        "ust2_low":  "MID_Low",
    })


def _read_wti():
    """WTI 시트 → wti (WTI 유가 종가=현재가)."""
    return _read_asof("WTI", {"wti": "현재가"})


def load(path=None):
    """LOAD_SPEC의 모든 엑셀을 읽어 sanity → ledger 적재(source='infomax')
    → 파생(fx_cum·roll_flag) 재계산 → CSV export. update.py와 동일 경로."""
    import store
    import sanity
    import update

    import config
    recs = []
    for spec in LOAD_SPEC:
        try:
            recs += _read_one(spec)
        except FileNotFoundError:
            print(f"  (건너뜀: 파일 없음 {spec['file']})")
        except KeyError:   # 워크북에 해당 시트 아직 없음(예: build 재실행 전) → 안전 스킵
            print(f"  (건너뜀: 시트 없음 '{spec['sheet']}' — build_infomax.py 재실행 필요)")
    recs += _read_macro()                         # MACRO 5종 (지수→yoy → 발표일 배치 → ffill)
    recs += _read_ust10()                         # 미국채10년(IR/US10Y MID_Close) → as-of 시프트
    recs += _read_ust2()                          # 미국채2년(IR/US02Y MID OHLC) → as-of 시프트
    recs += _read_wti()                           # WTI 유가(FRN/SPT:CL 현재가) → as-of 시프트
    ref = config.ref_date().isoformat()          # 미확정 당일 제외 (정산가 0 등)
    recs = [r for r in recs if r.date <= ref]

    conn = store.connect()
    store.init_db(conn)
    existing = {f: store.field_series(conn, f) for f in {r.field for r in recs}}
    passed, flagged, dropped = sanity.run(recs, existing)
    fetched = store.now_iso()
    store.upsert(conn, [(r.date, r.field, r.value, "infomax", r.as_of, fetched)
                        for r in passed])
    update.recompute_derived(conn)          # fx_cum(수급 누적) 등 갱신
    n = update.export_csv(conn)
    print(f"infomax load: 읽음 {len(recs)} / 적재 {len(passed)} "
          f"(플래그 {len(flagged)}, 격리 {len(dropped)})  → export {n}행")
    for r, why in (flagged + dropped)[:10]:
        print(f"  - {r.date} {r.field}={r.value} :: {why}")
    conn.close()


# 인포맥스 IMDH는 XLL 함수(_xll.IMDH) → 자동화 인스턴스엔 명시 등록 필요.
# 안 하면 #NAME?. (Excel 비트수에 맞는 XLL. IMX_XLL 환경변수로 오버라이드 가능)
IMX_XLL_32 = r"C:\Infomax\bin\excel\imxlexcelai.xll"
IMX_XLL_64 = r"C:\Infomax\bin\excel64\imxlexcelai64.xll"


def _register_addin(app):
    """인포맥스 XLL을 RegisterXLL로 등록해 IMDH를 사용 가능하게. Excel 비트수 자동판별."""
    import os
    override = os.environ.get("IMX_XLL")
    if override:
        cands = [override]
    else:
        is64 = "(x86)" not in (app.Path or "")
        cands = [IMX_XLL_64 if is64 else IMX_XLL_32, IMX_XLL_32, IMX_XLL_64]
    for x in cands:
        try:
            if os.path.exists(x) and app.RegisterXLL(x):
                return True
        except Exception:
            continue
    return False


def _com_retry(fn, tries=6, delay=2):
    """COM 'server busy'(0x800ac472) 등 일시적 실패를 재시도."""
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def _refresh_wb(app, path, wait):
    """워크북 하나: 종료일=기준일·개수 갱신 → 애드인 재계산 → 저장. (busy 재시도)"""
    import datetime as dt
    import os
    import time
    import config
    p = Path(path)
    if not p.exists():
        print(f"  (없음: {p.name})"); return
    # ★ 예전엔 F1=9000(26년치)라 열 때 IMDH 자동 재계산이 Excel을 바쁘게 만들어 이어지는 COM
    #   호출을 거부(RPC_E_CALL_REJECTED)했다. → (1)'수동 계산'으로 열고 명시적 재계산,
    #   (2)F1을 최근 N일(DAILY_REFRESH_COUNT)로 줄여 증분화 = 재계산량 자체를 축소.
    #   (IMDH는 count 기반·sort=D → 최근 N개만. 옛 데이터는 이미 ledger에 있고 upsert 멱등.)
    blank = app.Workbooks.Add() if app.Workbooks.Count == 0 else None
    try:
        app.Calculation = -4135        # xlCalculationManual
    except Exception:
        pass
    # blank 추가/계산모드 변경 직후 잠깐 바쁠 수 있음 → 유휴 대기 후 '직접' 열기.
    # (Open을 _com_retry로 감싸면 첫 시도가 거부됐어도 실제론 열려서, 재시도가 이중오픈→None 반환)
    _com_retry(lambda: app.Workbooks.Count, tries=12, delay=1)
    wb = app.Workbooks.Open(str(p.resolve()))
    try:
        ref = dt.datetime.combine(config.ref_date(), dt.time())
        count = int(os.environ.get("IMX_DAILY_COUNT", str(config.DAILY_REFRESH_COUNT)))
        for k in range(1, wb.Worksheets.Count + 1):              # 모든 시트 날짜창 갱신
            ws = wb.Worksheets(k)
            ws.Range("B1").Value = dt.datetime(2000, 1, 1)       # 시작일(형식상; IMDH는 무시)
            ws.Range("D1").Value = ref                            # 종료일 = 기준일(당일 마감후=당일)
            ws.Range("F1").Value = count                          # 최근 N거래일만(증분) — env로 조절
        _com_retry(lambda: app.CalculateFullRebuild())           # 명시적 재계산
        time.sleep(wait)
        _com_retry(lambda: app.CalculateFullRebuild())
        time.sleep(3)
        _com_retry(lambda: wb.Save())
        try:
            app.Calculation = -4105    # xlCalculationAutomatic (원복, 워크북 열려있을 때)
        except Exception:
            pass
        print(f"  refresh: {p.name}")
    finally:
        try:
            wb.Close(SaveChanges=True)
        except Exception:
            pass
        if blank is not None:
            try:
                blank.Close(SaveChanges=False)
            except Exception:
                pass


def refresh(path):
    """단일 파일 새로고침. Dispatch = 사용자 애드인(infomaxexcel.xlam) 로드됨.
    (DispatchEx는 애드인이 안 실려 #NAME? — 반드시 Dispatch)
    기존 Excel 있으면 attach(안 끔), 없으면 새로 시작(끝에 Quit). invisible."""
    import os
    import pythoncom
    import win32com.client as win32
    pythoncom.CoInitialize()
    try:
        app = win32.GetActiveObject("Excel.Application")   # 사용자 Excel 있으면 그걸로
        started = False
    except Exception:
        app = win32.Dispatch("Excel.Application")          # 없으면 새로(애드인 로드)
        started = True
    app.DisplayAlerts = False
    if started:
        app.Visible = False                                # 우리가 띄운 것만 숨김
    if not _register_addin(app):                           # ★ IMDH XLL 등록 (필수)
        print("  ⚠ 인포맥스 XLL 등록 실패 → IMDH #NAME? 위험 (IMX_XLL 확인)")
    try:
        _refresh_wb(app, path, int(os.environ.get("IMX_WAIT", "25")))
    finally:
        if started:
            try:
                app.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
    print("refresh 완료 →", path)


def _kill_headless_excel():
    """창 없는(자동화 잔재) Excel만 종료. 사용자 Excel(창 있음)은 보존."""
    import subprocess
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             "Get-Process EXCEL -ErrorAction SilentlyContinue | "
             "Where-Object {[string]::IsNullOrEmpty($_.MainWindowTitle)} | "
             "Stop-Process -Force -ErrorAction SilentlyContinue"],
            timeout=30, check=False,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


def ensure_sheets():
    """SHEETS 중 워크북에 없는 시트(신규 종목)를 openpyxl로 추가 — 엑셀/COM 불필요라 안전.
    수식만 심어두면 이어지는 refresh(CalculateFullRebuild)가 값을 채운다."""
    import datetime as dt
    import openpyxl
    import build_infomax as bi
    import config
    if not Path(IMX_FILE).exists():
        return []
    wb = openpyxl.load_workbook(IMX_FILE)
    existing = set(wb.sheetnames)
    start0 = dt.datetime(2000, 1, 1)
    ref = dt.datetime.combine(config.ref_date(), dt.time())
    added = []
    for name, mkt, sym, items in bi.SHEETS:
        if name in existing:
            continue
        ws = wb.create_sheet(name)
        for cell, val in [("A1", "시작"), ("B1", start0), ("C1", "종료"), ("D1", ref),
                          ("E1", "Data 개수"), ("F1", 9000), ("G1", "주기"), ("H1", "일"),
                          ("I1", "정렬"), ("J1", "D"), ("K1", "영업일"), ("L1", 0),
                          ("M1", "시세산출"), ("N1", "종가")]:
            ws[cell] = val
        for j, it in enumerate(items):
            ws.cell(3, j + 1, it)
        last = chr(ord("A") + len(items) - 1)
        # XLL 함수는 파일에 _xll. 접두로 저장됨(openpyxl은 자동변환 안 함) → 맞춰서 기입.
        ws["A2"] = (f'=_xll.IMDH("{mkt}","{sym}",A3:{last}3,$B$1,$D$1,$F$1,'
                    f'"Per="&$H$1&",sort="&$J$1&",real=false,Bizday="&$L$1&'
                    f'",Quote="&$N$1&",Pos=20,Orient=V,Title={name},DtFmt=1,TmFmt=1,unit=true")')
        added.append(name)
    if added:
        wb.save(IMX_FILE)
        print(f"  + 신규 시트 추가(openpyxl): {added}")
    return added


def refresh_all():
    """통합 파일(infomax_data.xlsx) 새로고침. 없는 시트(신규 종목)는 먼저 openpyxl로
    심고(ensure_sheets), COM refresh가 전 시트 재계산해 값을 채운다. 앞뒤로 잔재 Excel 청소."""
    _kill_headless_excel()
    if not Path(IMX_FILE).exists():
        print(f"  ⚠ {IMX_FILE} 없음 — 먼저 build_infomax.py로 생성 필요")
        return
    ensure_sheets()          # COM 아닌 openpyxl로 신규 시트 삽입(수식만) → refresh가 채움
    refresh(IMX_FILE)
    _kill_headless_excel()
    print("refresh_all 완료")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    mode = sys.argv[1]
    path = sys.argv[2] if len(sys.argv) > 2 else None
    if mode == "load":
        load()
    elif mode == "refresh":
        refresh(path) if path else refresh_all()   # 경로 없으면 통합 파일 전체
    elif mode == "dump":
        dump(path)
    else:
        print(__doc__)
