# -*- coding: utf-8 -*-
"""
infomax_pull.py — 인포맥스 애드인 엑셀 → ledger 리더/수집기

인포맥스 애드인(IMDH 히스토리 함수)으로 채운 엑셀을 읽어 ledger에 적재한다.
애드인은 그 PC의 Excel + 인포맥스 로그인 세션에서만 값이 산다.

데이터는 통합 워크북 infomax_data.xlsx(시트 7개, 필요한 항목만). build_infomax.py로 생성.

모드:
  dump    "경로.xlsx"  — 엑셀 시트/셀 레이아웃 출력 (캐시값, Excel 불필요)
  refresh [경로]       — COM 재조회. 경로 없으면 infomax_data.xlsx(시트 전체) 갱신
  load                 — infomax_data.xlsx → sanity → ledger 적재 → 파생계산 → export

일일 자동화 순서: refresh → load  (run.py가 update.py와 함께 호출)
LOAD_SPEC = 인포맥스 시트·항목 매핑 / build_infomax.py의 SHEETS = 종목 정의(단일 진실).
"""
import sys
from pathlib import Path


def dump(path):
    """저장된 엑셀의 모든 시트 used range를 출력. IMDH 캐시값을 그대로 보여준다."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)   # data_only = 캐시된 계산값
    print(f"파일: {path}")
    print(f"시트: {wb.sheetnames}\n")
    for ws in wb.worksheets:
        print("=" * 60)
        print(f"[{ws.title}]  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}")
        # 앞 12행만 미리보기 (레이아웃 파악엔 충분)
        for r, row in enumerate(ws.iter_rows(values_only=True), 1):
            if r > 12:
                print("   ... (이하 생략)")
                break
            cells = ["" if c is None else str(c) for c in row]
            while cells and cells[-1] == "":
                cells.pop()
            if cells:
                print(f"  r{r:<3} " + " | ".join(cells))


# ── 읽기 스펙: 생성된 엑셀 → 우리 field 매핑 ──────────────────────────
# IMDH 출력: '일자' 헤더행 아래로 데이터. 값 = 지정한 '항목 헤더 텍스트'의 열.
# 통합 워크북 infomax_data.xlsx (시트별 IMDH, 필요한 항목만). build_infomax.py로 생성.
# 각 항목: sheet, header_text(날짜열=일자, row3), fields{우리필드: 항목헤더 or ("net",매수,매도)}
# 프로젝트 루트(=src/의 부모) 기준 절대경로 → 실행 cwd와 무관하게 워크북을 찾음.
IMX_FILE = str(Path(__file__).resolve().parent.parent / "infomax_data.xlsx")


def _fut_fields(suf):
    """FUT 시트(연결선물) → 우리 필드 매핑. 순매수 = 매수수량−매도수량."""
    return {
        f"ktb{suf}_settle": "현재가",           # 정산가는 옛 데이터 0 많음 → 현재가(종가)
        f"ktb{suf}_chg": "전일대비",
        f"ktb{suf}_open": "시가",
        f"ktb{suf}_high": "고가",
        f"ktb{suf}_low": "저가",
        f"ktb{suf}_vol": "거래량",
        f"ktb{suf}_oi": "미결제약정수량",
        f"ktb{suf}_theo_basis": "이론베이시스",
        f"fx_net_ktb{suf}": ("net", "외국인합매수수량", "외국인합매도수량"),
        f"bank_net_ktb{suf}": ("net", "은행매수수량", "은행매도수량"),
        f"itrust_net_ktb{suf}": ("net", "투신매수수량", "투신매도수량"),
        f"ins_net_ktb{suf}": ("net", "보험매수수량", "보험매도수량"),
    }


LOAD_SPEC = [
    {"sheet": "FUT3",   "header_text": "일자", "fields": _fut_fields("3")},
    {"sheet": "FUT10",  "header_text": "일자", "fields": _fut_fields("10")},
    {"sheet": "IRS3Y",  "header_text": "일자", "fields": {"irs3y": "MID종가"}},
    {"sheet": "IRS10Y", "header_text": "일자", "fields": {"irs10y": "MID종가"}},
    {"sheet": "OIS3Y",  "header_text": "일자", "fields": {"ois3y": "MID종가"}},
    {"sheet": "KOFR",   "header_text": "일자", "fields": {"kofr": "현재가"}},  # ECO/785831 (ECOS보다 신선)
]
for _s in LOAD_SPEC:            # 전부 같은 통합 파일
    _s["file"] = IMX_FILE


def _to_date(v):
    import datetime as dt
    if isinstance(v, dt.datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()[:10].replace("/", "-").replace(".", "-")
    return s if len(s) == 10 else None


def _read_one(spec):
    """LOAD_SPEC 항목 하나 → Record 리스트 (헤더 텍스트로 열 매칭)."""
    import openpyxl
    from collectors.base import Record
    ws = openpyxl.load_workbook(spec["file"], data_only=True)[spec["sheet"]]
    rows = list(ws.iter_rows(values_only=True))
    hi = next(i for i, r in enumerate(rows)
              if r and spec["header_text"] in [str(c).strip() if c else "" for c in r])
    header = [str(c).strip() if c else "" for c in rows[hi]]
    dcol = header.index(spec["header_text"])

    def cell(r, name):
        c = header.index(name)
        return r[c] if c < len(r) else None

    recs = []
    for our_field, item in spec["fields"].items():
        for r in rows[hi + 1:]:
            d = _to_date(r[dcol]) if r[dcol] is not None else None
            if not d:
                continue
            if isinstance(item, tuple) and item[0] == "net":   # 매수 − 매도
                buy, sell = cell(r, item[1]), cell(r, item[2])
                if buy in (None, "") or sell in (None, ""):
                    continue
                v = float(buy) - float(sell)
            else:                                              # 직접 열
                raw = cell(r, item)
                if raw in (None, ""):
                    continue
                v = float(raw)
            # 가격필드(settle/OHLC/basis) 정확히 0 = 미기록 → 빈칸(0=진짜 규칙).
            # vol/oi/net의 0은 진짜(거래없음/순매수0)라 유지.
            if v == 0 and our_field.endswith(
                    ("_settle", "_open", "_high", "_low", "_theo_basis")):
                continue
            recs.append(Record(date=d, field=our_field, value=v, as_of=d))
    return recs


def load(path=None):
    """LOAD_SPEC의 모든 엑셀을 읽어 sanity → ledger 적재(source='infomax')
    → 파생(fx_cum·roll_flag) 재계산 → CSV export. update.py와 동일 경로."""
    import store
    import sanity
    import update

    import config
    recs = []
    for spec in LOAD_SPEC:
        try:
            recs += _read_one(spec)
        except FileNotFoundError:
            print(f"  (건너뜀: 파일 없음 {spec['file']})")
        except KeyError:   # 워크북에 해당 시트 아직 없음(예: build 재실행 전) → 안전 스킵
            print(f"  (건너뜀: 시트 없음 '{spec['sheet']}' — build_infomax.py 재실행 필요)")
    ref = config.ref_date().isoformat()          # 미확정 당일 제외 (정산가 0 등)
    recs = [r for r in recs if r.date <= ref]

    conn = store.connect()
    store.init_db(conn)
    existing = {f: store.field_series(conn, f) for f in {r.field for r in recs}}
    passed, flagged, dropped = sanity.run(recs, existing)
    fetched = store.now_iso()
    store.upsert(conn, [(r.date, r.field, r.value, "infomax", r.as_of, fetched)
                        for r in passed])
    update.recompute_derived(conn)          # fx_cum(수급 누적) 등 갱신
    n = update.export_csv(conn)
    print(f"infomax load: 읽음 {len(recs)} / 적재 {len(passed)} "
          f"(플래그 {len(flagged)}, 격리 {len(dropped)})  → export {n}행")
    for r, why in (flagged + dropped)[:10]:
        print(f"  - {r.date} {r.field}={r.value} :: {why}")
    conn.close()


# 인포맥스 IMDH는 XLL 함수(_xll.IMDH) → 자동화 인스턴스엔 명시 등록 필요.
# 안 하면 #NAME?. (Excel 비트수에 맞는 XLL. IMX_XLL 환경변수로 오버라이드 가능)
IMX_XLL_32 = r"C:\Infomax\bin\excel\imxlexcelai.xll"
IMX_XLL_64 = r"C:\Infomax\bin\excel64\imxlexcelai64.xll"


def _register_addin(app):
    """인포맥스 XLL을 RegisterXLL로 등록해 IMDH를 사용 가능하게. Excel 비트수 자동판별."""
    import os
    override = os.environ.get("IMX_XLL")
    if override:
        cands = [override]
    else:
        is64 = "(x86)" not in (app.Path or "")
        cands = [IMX_XLL_64 if is64 else IMX_XLL_32, IMX_XLL_32, IMX_XLL_64]
    for x in cands:
        try:
            if os.path.exists(x) and app.RegisterXLL(x):
                return True
        except Exception:
            continue
    return False


def _com_retry(fn, tries=6, delay=2):
    """COM 'server busy'(0x800ac472) 등 일시적 실패를 재시도."""
    import time
    last = None
    for _ in range(tries):
        try:
            return fn()
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


def _refresh_wb(app, path, wait):
    """워크북 하나: 종료일=기준일·개수 갱신 → 애드인 재계산 → 저장. (busy 재시도)"""
    import datetime as dt
    import time
    import config
    p = Path(path)
    if not p.exists():
        print(f"  (없음: {p.name})"); return
    wb = _com_retry(lambda: app.Workbooks.Open(str(p.resolve())))
    try:
        ref = dt.datetime.combine(config.ref_date(), dt.time())
        for k in range(1, wb.Worksheets.Count + 1):              # 모든 시트 날짜창 갱신
            ws = wb.Worksheets(k)
            ws.Range("B1").Value = dt.datetime(2000, 1, 1)       # 시작일 = 가용 최대
            ws.Range("D1").Value = ref                            # 종료일 = 기준일
            ws.Range("F1").Value = 9000                           # 개수 (전체 이력)
        _com_retry(lambda: app.CalculateFullRebuild())
        time.sleep(wait)
        _com_retry(lambda: app.CalculateFullRebuild())
        time.sleep(3)
        _com_retry(lambda: wb.Save())
        print(f"  refresh: {p.name}")
    finally:
        try:
            wb.Close(SaveChanges=True)
        except Exception:
            pass


def refresh(path):
    """단일 파일 새로고침. Dispatch = 사용자 애드인(infomaxexcel.xlam) 로드됨.
    (DispatchEx는 애드인이 안 실려 #NAME? — 반드시 Dispatch)
    기존 Excel 있으면 attach(안 끔), 없으면 새로 시작(끝에 Quit). invisible."""
    import os
    import pythoncom
    import win32com.client as win32
    pythoncom.CoInitialize()
    try:
        app = win32.GetActiveObject("Excel.Application")   # 사용자 Excel 있으면 그걸로
        started = False
    except Exception:
        app = win32.Dispatch("Excel.Application")          # 없으면 새로(애드인 로드)
        started = True
    app.DisplayAlerts = False
    if started:
        app.Visible = False                                # 우리가 띄운 것만 숨김
    if not _register_addin(app):                           # ★ IMDH XLL 등록 (필수)
        print("  ⚠ 인포맥스 XLL 등록 실패 → IMDH #NAME? 위험 (IMX_XLL 확인)")
    try:
        _refresh_wb(app, path, int(os.environ.get("IMX_WAIT", "25")))
    finally:
        if started:
            try:
                app.Quit()
            except Exception:
                pass
        try:
            pythoncom.CoUninitialize()
        except Exception:
            pass
    print("refresh 완료 →", path)


def _kill_headless_excel():
    """창 없는(자동화 잔재) Excel만 종료. 사용자 Excel(창 있음)은 보존."""
    import subprocess
    try:
        subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             "Get-Process EXCEL -ErrorAction SilentlyContinue | "
             "Where-Object {[string]::IsNullOrEmpty($_.MainWindowTitle)} | "
             "Stop-Process -Force -ErrorAction SilentlyContinue"],
            timeout=30, check=False,
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    except Exception:
        pass


def refresh_all():
    """통합 파일(infomax_data.xlsx) 하나만 새로고침(시트 7개 한 번에 재계산).
    시작/끝에 자동화 잔재 Excel 청소."""
    _kill_headless_excel()
    if not Path(IMX_FILE).exists():
        print(f"  ⚠ {IMX_FILE} 없음 — 먼저 build_infomax.py로 생성 필요")
        return
    refresh(IMX_FILE)
    _kill_headless_excel()
    print("refresh_all 완료")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)
    mode = sys.argv[1]
    path = sys.argv[2] if len(sys.argv) > 2 else None
    if mode == "load":
        load()
    elif mode == "refresh":
        refresh(path) if path else refresh_all()   # 경로 없으면 통합 파일 전체
    elif mode == "dump":
        dump(path)
    else:
        print(__doc__)
